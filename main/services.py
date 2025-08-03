import pandas as pd
from settings import *
import os
from watch import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ! = IMPORTANTE
# ? = OQUE FAZ
# ^ = A FAZER
# * = AVISOS


class Geral:

    def __init__(self, PathXlsx):
        self.Path = PathXlsx

    def MakeArq(self):
        if self.Path.endswith(".xlsx"):
            self.csv = pd.read_excel(self.Path)
        elif self.Path.endswith(".csv"):
            self.csv = pd.read_csv(self.Path, sep=";")
        self.csv.to_csv(
            os.path.join(TEMP, "Arq.csv"), index=True, encoding="utf-8", sep=";"
        )

    def Tratamento(self):
        lista = {}
        self.Geral = pd.DataFrame()
        for i, rows in self.csv.iterrows():
            lista["cod"] = rows["cod"]
            lista["nome"] = rows["nome"]
            lista["cpf"] = rows["cpf"]
            self.Geral = pd.concat(
                [self.Geral, pd.DataFrame([lista])], ignore_index=False
            )
        self.Geral.to_excel(os.path.join(FINAL, "DeuCerto.xlsx"), index=False)

    def LGPD(self):
        conteudo = os.listdir(f"{os.path.join(TEMP)}")
        arquivos = [item for item in conteudo]
        for i in arquivos:
            os.remove(f"{os.path.join(TEMP, i)}")

    def Espaco(self):
        wb = load_workbook(os.path.join(FINAL, "DeuCerto.xlsx"))
        sheet = wb.sheetnames[0]
        ws = wb[sheet]
        for col in ws.columns:
            max_l = 0
            coluna = col[0].column
            coluna_letra = get_column_letter(coluna)
            for cell in col:
                try:
                    if cell.value:
                        max_l = max(max_l, len(str(cell.value)))
                except Exception as e:
                    pass
            ajuste = max_l + 2
            ws.column_dimensions[coluna_letra].width = ajuste
        wb.save(rf"{FINAL}\DeuCerto.xlsx")

    # ? Pega index de coluna especifica
    # ? Exemplo: print(self.csv.columns.to_list()) vai voltar uma lista com todas as colunas do arquivo
    # ? ['cod', 'nome', 'cpf', 'telefone', 'cc', 'valor', 'local']
    # ? Vamos dizer que search = 'cc'
    # ? index = self.csv.columns.to_list().index(search)
    # ? print(index) >>> vai retornar 4
    # ? return alfabeto[index] >>> vai retornar 'e'
    def TakeIndex(self, search):
        alfabeto = [
            "a",
            "b",
            "c",
            "d",
            "e",
            "f",
            "g",
            "h",
            "i",
            "j",
            "k",
            "l",
            "m",
            "n",
            "o",
            "p",
            "q",
            "r",
            "s",
            "t",
            "u",
            "v",
            "w",
            "x",
            "y",
            "z",
        ]
        index = self.csv.columns.to_list().index(search)
        return alfabeto[index]


class Financeiro(Geral):

    def __init__(self, xlsx, csv):
        super().__init__(PathCsv=csv, PathXlsx=xlsx)

    def CentroCustos(self):
        centro = self.Centros()
        lista = []
        for i in range(0, len(centro)):
            cut = str(centro[i])
            intercompany = f"{cut[0]}{cut[1]}"
            CentroCusto = f"{cut[2]}{cut[3]}"
            if intercompany == "81":
                emp = "TVCA"
            if intercompany == "65":
                emp = "Portal MT"
            if intercompany == "69":
                emp = "FMCA"
            if intercompany == "85":
                emp = "On Line(MT)"
            if CentroCusto == "03":
                area = "ADM"
            if CentroCusto == "06":
                area = "RH"
            if CentroCusto == "10":
                area = "COMERCIA"
            if CentroCusto == "11":
                area = "OPEC"
            if CentroCusto == "12":
                area = "MKT"
            if CentroCusto == "14":
                area = "PROGRAMAÇÃO"
            if CentroCusto == "15":
                area = "JORNALISMO"
            if CentroCusto == "21":
                area = "TECNOLOGIA"
            # TODO ADICIONAR AREA DA DIRETORIA !
        lista.append(f"{centro[i]} - {area} - {emp}")
        return lista

    def Limpa(self, lista):
        return list(dict.fromkeys(sorted(lista)))

    def Centros(self):
        lista = self.Limpa((self.csv.cc.to_list()))
        return lista

    # ? Função para soma

    # ? def Sum(self):
    # ?     lis = []
    # ?     for i in range(1, 6):
    # ?         num = f'=SUM(A{2 + i - 1}:D{2 + i - 1})'
    # ?         lis.append(num)
    # ?     self.csv['tbl5'] = lis
    # ?     self.csv.to_csv(self.pathCSV, index=True)
    # ?     self.csv.to_excel('arquivos/Pasta1.xlsx', index=False)
